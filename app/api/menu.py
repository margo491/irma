from io import BytesIO
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.database import get_db
from app.models.menu import MenuCategory, MenuItem
from app.schemas.menu import CategoryOut, MenuItemOut

router = APIRouter()
items_router = APIRouter()

SITE_URL = "https://irma-cafe.ru"


@router.get("/", response_model=list[CategoryOut])
async def get_categories(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(MenuCategory).order_by(MenuCategory.sort_order))
    return result.scalars().all()


@router.get("/export.yml")
async def export_menu_yml(db: AsyncSession = Depends(get_db)):
    """YML (Яндекс.Маркет) фид меню — для импорта каталога в ChatMarket по URL.

    ChatMarket поддерживает импорт из XML-фида по ссылке; точная схема не
    задокументирована публично, YML — самый распространённый формат таких
    фидов на российском рынке, используем его как отправную точку."""
    categories = (
        (await db.execute(select(MenuCategory).order_by(MenuCategory.sort_order)))
        .scalars()
        .all()
    )
    items = (
        (await db.execute(select(MenuItem).order_by(MenuItem.category_id)))
        .scalars()
        .all()
    )

    root = Element("yml_catalog", {"date": datetime.utcnow().strftime("%Y-%m-%d %H:%M")})
    shop = SubElement(root, "shop")
    SubElement(shop, "name").text = "IrMa"
    SubElement(shop, "company").text = "Семейное кафе IrMa by Marinich"
    SubElement(shop, "url").text = SITE_URL

    currencies = SubElement(shop, "currencies")
    SubElement(currencies, "currency", {"id": "RUR", "rate": "1"})

    categories_el = SubElement(shop, "categories")
    for cat in categories:
        SubElement(categories_el, "category", {"id": str(cat.id)}).text = cat.name

    offers_el = SubElement(shop, "offers")
    for item in items:
        offer = SubElement(
            offers_el,
            "offer",
            {"id": str(item.id), "available": "true" if item.is_available else "false"},
        )
        SubElement(offer, "name").text = item.name
        SubElement(offer, "price").text = str(item.price)
        SubElement(offer, "currencyId").text = "RUR"
        SubElement(offer, "categoryId").text = str(item.category_id)
        if item.image_url:
            url = item.image_url
            if not url.startswith("http"):
                url = f"{SITE_URL}/{url.lstrip('/')}"
            SubElement(offer, "picture").text = url
        if item.description:
            SubElement(offer, "description").text = item.description

    xml_bytes = tostring(root, encoding="utf-8")
    pretty = minidom.parseString(xml_bytes).toprettyxml(indent="  ", encoding="utf-8")
    return Response(content=pretty, media_type="application/xml")


@router.get("/export.xlsx")
async def export_menu_xlsx(db: AsyncSession = Depends(get_db)):
    """Excel-выгрузка меню под реальный шаблон импорта ChatMarket
    (5 листов: Категории, Товары, Характеристики, Дополнительные услуги,
    URL-кнопки — структура подтверждена их собственным example.xlsx).

    Категории и Товары заполняются из нашего меню. Остальные три листа
    для кафе неприменимы (размеры/варианты, платные допуслуги, кнопки со
    ссылками — это под магазины одежды/техники) — оставлены с одной
    строкой заголовков, без выдуманных данных."""
    categories = (
        (await db.execute(select(MenuCategory).order_by(MenuCategory.sort_order)))
        .scalars()
        .all()
    )
    items = (
        (await db.execute(select(MenuItem).order_by(MenuItem.category_id)))
        .scalars()
        .all()
    )

    wb = Workbook()

    ws_cat = wb.active
    ws_cat.title = "Категории"
    ws_cat.append(["ID категории", "Название"])
    for cat in categories:
        ws_cat.append([f"C{cat.id}", cat.name])

    ws_items = wb.create_sheet("Товары")
    ws_items.append([
        "ID категории", "ID товара", "Название", "Цена", "Скидка", "Бонусов",
        "Описание", "Количество", "Артикул",
        *[f"Изображение {i}" for i in range(1, 11)],
    ])
    for item in items:
        image_url = item.image_url or None
        if image_url and not image_url.startswith("http"):
            image_url = f"{SITE_URL}/{image_url.lstrip('/')}"
        row = [
            f"C{item.category_id}",
            f"I{item.id}",
            item.name,
            float(item.price),
            0,
            None,
            item.description or "",
            None,
            None,
            image_url,
        ] + [None] * 9  # Изображение 2..10
        ws_items.append(row)

    ws_char = wb.create_sheet("Характеристики")
    ws_char.append(["ID товара", "Название", "Значение"])

    ws_extra = wb.create_sheet("Дополнительные услуги")
    ws_extra.append(["ID товара", "Название", "Цена"])

    ws_url = wb.create_sheet("URL-кнопки")
    ws_url.append(["ID товара", "Текст", "Ссылка"])

    for ws in (ws_cat, ws_items, ws_char, ws_extra, ws_url):
        for i in range(1, ws.max_column + 1):
            col = get_column_letter(i)
            values = [ws.cell(row=1, column=i).value] + [
                ws.cell(row=r, column=i).value for r in range(2, ws.max_row + 1)
            ]
            max_len = max(len(str(v)) for v in values if v is not None) if any(v is not None for v in values) else 10
            ws.column_dimensions[col].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=irma-menu.xlsx"},
    )


@router.get("/{category_id}", response_model=list[MenuItemOut])
async def get_items_by_category(category_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(MenuItem).where(MenuItem.category_id == category_id, MenuItem.is_available == True)
    )
    return result.scalars().all()


@items_router.get("/{item_id}", response_model=MenuItemOut)
async def get_item(item_id: int, db: AsyncSession = Depends(get_db)):
    item = await db.get(MenuItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Позиция не найдена")
    return item
